from collections import defaultdict

import openpyxl
import re
import os
import datetime


def parseWork(path):
    wb = openpyxl.load_workbook(path)
    all_sheets = wb.sheetnames
    for i in range(len(all_sheets)):
        sheet = wb[all_sheets[i]]
        # print(sheet.title + ': max_row: ' + str(sheet.max_row) + '  max_column: ' + str(sheet.max_column))

        col, row = getWorkByPerson(sheet)
        while row < sheet.max_row:
            row = row + 1
            if sheet.cell(row, col).value and sheet.cell(row, col + 2).value:
                yield sheet.cell(row, col).value, '\n'.join(
                    [re.search(r"方正项目周报（([\w-]+至[\w-]+)）", path).group(1) + ':',
                     re.sub(r"\n[\s| ]*\n", '\n', sheet.cell(row, col + 2).value)])


def getWorkByPerson(sheet):
    for column in sheet.iter_cols():
        for cell2 in column:
            if cell2.value is not None:
                info2 = cell2.value.find('支撑人员')
                if info2 == 0:
                    row, col = cell2.row, cell2.column
                    return col, row


def getWorks():
    regex = r"方正项目周报（((2019|2020)-(10|11|12)-[0-9]{1,2})至"
    fileList = [item for item in os.listdir('.') if re.match(regex, item)]
    # for k in fileList:
    #     print(re.search(regex, k).groups())
    fileList.sort(key=lambda k: datetime.datetime.strptime(re.search(regex, k).group(1), '%Y-%m-%d'))
    print(fileList)

    result = defaultdict(str)
    for work in fileList:
        for key, val in parseWork(work):
            result[key] = '\n'.join([result[key], val])
    return result


os.chdir('F:\\document')
with open('data.txt ', 'w') as f:
    for item, value in getWorks().items():
        f.write('\n' + (item + value).replace('\xa0', ' '))
